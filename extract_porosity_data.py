from PIL import Image
import os
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import time

# Loops through the scans pixel by pixel using PIL.
# Extracts the color information each pixel.
# With this information, the porosity percentage of each scan is calculated.

# Path of the folder, which contains the CT-Scans of all the samples 
path = "C:/Users/User/Desktop/CT Scans Final/"

# Opens Excel file
wb = load_workbook("porosity_data_sample_1.xlsx")
ws = wb.active

# Variables for porosity calculation
red_pixel_count = 0
colored_pixel_count = 0
black_pixel_count = 0
counter_element = 0


# Loops the samples
for sample in os.listdir(path):

    # Loops the 15 elements of each sample
    for element in os.listdir(path + sample):
        counter_element = counter_element + 1
        # Loops the 3 surfaces of each element
        for surface in os.listdir(path + sample + "/" + element):
            
            # Ignores the files not containing the relevant scans
            if "Ebene" in surface:
                
                # Loops the scans of each surface
                for scan in os.listdir(path + sample + "/" + element + "/" + surface):
                    
                    # Checks for .jpg file
                    if scan.endswith(".jpg"):

                        # Opens the current scan
                        image = Image.open(path + sample + "/" + element + "/" + surface + "/" + scan)
                        img_width, img_height = image.size
                        img_pixel = image.load()

                        start = time.perf_counter()

                        # Loops the individual pixels
                        for j in range (0, img_height):
                            for i in range(0, img_width):
                                if img_pixel[i ,j] != (0, 0, 0):
                                    colored_pixel_count = colored_pixel_count + 1
                                if img_pixel[i, j] == (254, 0, 0):
                                    red_pixel_count = red_pixel_count + 1
                                if img_pixel[i, j] == (0, 0, 0):
                                    black_pixel_count = black_pixel_count +1

                        end = time.perf_counter()
                        time_passed = round((end - start), 2)
                        # Add timer print() func

                        # Calculates porosity   
                        # img_pixel_total = img_height * img_width
                        porosity =((black_pixel_count / (colored_pixel_count - red_pixel_count))) * 100

                        ws["A" + str(counter_element)] = element
                        wb.save("porosity_data_sample_1.xlsx")
                        print(str(porosity) + " % " + scan)