import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import itertools

wb = load_workbook("porosity_data_sample_1_1.xlsx")
ws = wb.active

ws.unmerge_cells("A1:A3")
ws.delete_cols(1, 2)

scan_count = len([scan for scan in os.listdir("C:/Users/User/Desktop/CT Scans Final/1/1 1/Ebene 1/")])

ws.insert_rows(1)
ws.insert_cols(1)

ws["A2"].value = "Ebene 1"
ws["A3"].value = "Ebene 2"
ws["A4"].value = "Ebene 3"

a = 0

for cell in ws[1]:
    cell.value = a
    a = a + 1
    if a == scan_count + 1:
        break

ws["A1"].value = None

wb.save("delete_column.xlsx")

df = pd.read_excel("df/delete_column.xlsx")

print(df)