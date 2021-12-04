import os
from PIL import Image
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

file = filedialog.askopenfilename()

wb = load_workbook(file)
ws = wb.active


for row in ws.iter_rows(min_row=2, max_row=2, max_col=4):
    for cell in row:
        print(cell.value)



wb.save(file)






'''
path = "C:/Users/User/Desktop/CT Scans Final/1/1 1/Ebene 1/"


wb = load_workbook("porosity_data_sample_1_backup.xlsx")
ws = wb.active

a = 0

for row in ws.iter_rows(min_row=2, max_row=2, max_col=4):
    for cell in row:
        print(cell.value)

b = len([scan for scan in os.listdir(path)])

print(b)

wb.save("porosity_data_sample_1_backup.xlsx")
'''