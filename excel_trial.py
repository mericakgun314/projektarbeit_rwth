from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
from PIL import Image

# Path of the folder, which contains the CT-Scans of all the samples 
path = "C:/Users/User/Desktop/CT Scans Final/1/1 1/"

wb = load_workbook("porosity_data_sample_1.xlsx")
ws = wb.active

red_pixel_count = 0
colored_pixel_count = 0
black_pixel_count = 0
counter_element = 0

for ebene in os.listdir(path):

    if "Ebene" in ebene:
        ws["A1"] = str(path.split("/")[6])
        ws.merge_cells("A1:A3")
        ws["A1"].alignment = Alignment(horizontal='center', vertical='center')

        ws["B" + str(ebene.split()[1])] = str(ebene)
        wb.save("porosity_data_sample_1.xlsx")

        for scan in os.listdir(path + ebene):
            
            if scan.endswith(".jpg"):
                
                # Opens the current scan
                full_path = path + ebene + "/" + scan

                split_scan_path = scan.split()
                
                image = Image.open(full_path)
                img_width, img_height = image.size
                img_pixel = image.load()

                # Loops the individual pixels
                for j in range (0, img_height):
                    for i in range(0, img_width):
                        if img_pixel[i ,j] != (0, 0, 0):
                            colored_pixel_count = colored_pixel_count + 1
                        if img_pixel[i, j] == (254, 0, 0):
                            red_pixel_count = red_pixel_count + 1
                        if img_pixel[i, j] == (0, 0, 0):
                            black_pixel_count = black_pixel_count +1
                        
                # Calculates porosity   
                img_pixel_total = img_height * img_width
                porosity =((black_pixel_count / (colored_pixel_count - red_pixel_count))) * 100

                # wb.save("porosity_data_sample_1.xlsx")
                print(str(porosity) + " % " + scan)