from types import NoneType
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import filedialog

root = tk.Tk()
root.withdraw()

path = filedialog.askdirectory() + "/"
i = 0

for file in os.listdir(path):
    if file.endswith(".xlsx"):

        wb = load_workbook(path + file)
        ws = wb.active

        i = i + 1
        for row in ws.iter_rows(min_row=1, min_col=1, max_col=8):
            for cell in row:
                
                if "x" in str(cell.value):
                    split_cell = cell.value.split("x")
                    placement = round(float(split_cell[0]) + float(split_cell[1]), 3)    
                    cell.value = str(placement)

        wb.save("textile_placement_" + str(i) + ".xlsx")            