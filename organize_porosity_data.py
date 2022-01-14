from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import filedialog

# Makes adjustments to the files in the folders "raw_sample_x", the values remain unchanged.

# Organizes the Excel files, which are created by "porosity_to_excel.py".
# Adds columns, headers and saves these Excel files as new and relocates them into their respective folder.

root = tk.Tk()
root.withdraw()

path = filedialog.askdirectory() + "/"

for file in os.listdir(path):
    if file.endswith(".xlsx"):
        wb = load_workbook(path + file)
        ws = wb.active

        ws.unmerge_cells("A1:A3")
        ws.delete_cols(1, 2)

        file_split = file.split("_")

        scan_count = len([scan for scan in os.listdir("C:/Users/User/Desktop/CT Scans Final/" + file_split[3] + "/" + file_split[3] + " " + file_split[4].split(".")[0] + "/" + "Ebene 1")])

        ws.insert_rows(1)
        ws.insert_cols(1)

        ws["A2"].value = "Ebene 1"
        ws["A3"].value = "Ebene 2"
        ws["A4"].value = "Ebene 3"

        a = 0

        for cell in ws[1]:
            cell.value = a
            a = a + 1
            if a == scan_count + 1:
                break

        ws["A1"].value = None

        for row in ws.iter_rows(min_row=2, max_row=4, min_col=2):
            for cell in row:
                if cell.value != None:
                    split_cell = cell.value.split(" ")
                    decimal = split_cell[0].replace(".", ",")
                    cell.value = decimal

        wb.save("df_" + file_split[3] + "/df_" + file_split[3] + "_" + file_split[4])