from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os
from PIL import Image
import tkinter as tk
from tkinter import filedialog

# Loops through the scans pixel by pixel using PIL.
# Extracts the color information each pixel.
# With this information, the porosity percentage of each scan is calculated.
# Saves these information to Excel files for each element.

# Prompts the file explorer, which contains the scans
root = tk.Tk()
root.withdraw()

def surface_selection():
    print("Select a surface folder:")
    path = filedialog.askdirectory() + "/"
    ebene = path.split("/")[7]

    print("Select an existing excel file:")
    excel_file = filedialog.askopenfilename()
    
    wb = load_workbook(excel_file)
    ws = wb.active

    if "Ebene" in ebene:
            # Creates the headlines for the table
            ws["A1"] = str(path.split("/")[6])
            ws.merge_cells("A1:A3")
            ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
            ws["B" + str(ebene.split()[1])] = str(ebene)

            counter = 0

            for scan in os.listdir(path):
                
                if scan.endswith(".jpg"):
                    
                    # Opens the current scan
                    full_path = path + "/" + scan
                    
                    image = Image.open(full_path)
                    img_width, img_height = image.size
                    img_pixel = image.load()

                    red_pixel_count = 0
                    colored_pixel_count = 0
                    black_pixel_count = 0

                    # Loops the individual pixels
                    for j in range (0, img_height):
                        for i in range(0, img_width):
                            if img_pixel[i ,j] != (0, 0, 0):
                                colored_pixel_count = colored_pixel_count + 1
                            if img_pixel[i, j] == (254, 0, 0):
                                red_pixel_count = red_pixel_count + 1
                            if img_pixel[i, j] == (0, 0, 0):
                                black_pixel_count = black_pixel_count +1
                    
                    # Calculates porosity   
                    img_pixel_total = img_height * img_width
                    porosity =((black_pixel_count / (colored_pixel_count - red_pixel_count))) * 100
                    
                    # Gets the number of scans
                    scan_count = len([scan for scan in os.listdir(path)])
                    
                    # Writes porosity data
                    for row in ws.iter_rows(min_row=int(ebene.split(" ")[1]), max_row=int(ebene.split(" ")[1]), min_col=2, max_col= (2 + scan_count)):
                        for cell in row:
                            if cell.value == None:
                                cell.value = str(round(porosity, 4)) + " %"
                                counter = counter + 1
                                break
                    
                    wb.save(excel_file)

                    print("Progress: " + str(counter) + " / " + str(scan_count) + ", " + str(round(((counter / scan_count) * 100), 1)) + " %")


def element_selection():
    path = filedialog.askdirectory() + "/"

    # Initializes Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active

    for ebene in os.listdir(path):
        
        if "Ebene" in ebene:
            # Creates the headlines for the table
            ws["A1"] = str(path.split("/")[6])
            ws.merge_cells("A1:A3")
            ws["A1"].alignment = Alignment(horizontal='center', vertical='center')
            ws["B" + str(ebene.split()[1])] = str(ebene)

            counter = 0

            for scan in os.listdir(path + ebene):
                
                if scan.endswith(".jpg"):
                    
                    # Opens the current scan
                    full_path = path + ebene + "/" + scan
                    
                    image = Image.open(full_path)
                    img_width, img_height = image.size
                    img_pixel = image.load()

                    red_pixel_count = 0
                    colored_pixel_count = 0
                    black_pixel_count = 0

                    # Loops the individual pixels
                    for j in range (0, img_height):
                        for i in range(0, img_width):
                            if img_pixel[i ,j] != (0, 0, 0):
                                colored_pixel_count = colored_pixel_count + 1
                            if img_pixel[i, j] == (254, 0, 0):
                                red_pixel_count = red_pixel_count + 1
                            if img_pixel[i, j] == (0, 0, 0):
                                black_pixel_count = black_pixel_count +1
                    
                    # Calculates porosity   
                    img_pixel_total = img_height * img_width
                    porosity =((black_pixel_count / (colored_pixel_count - red_pixel_count))) * 100
                    
                    # Gets the number of scans
                    scan_count = len([scan for scan in os.listdir(path + ebene)])
                    
                    # Writes porosity data
                    for row in ws.iter_rows(min_row=int(ebene.split(" ")[1]), max_row=int(ebene.split(" ")[1]), min_col=2, max_col= (2 + scan_count)):
                        for cell in row:
                            if cell.value == None:
                                cell.value = str(round(porosity, 4)) + " %"
                                counter = counter + 1
                                break
                    
                    wb.save("porosity_data_sample_" + str(path.split("/")[6]).split(" ")[0] + "_" + str(path.split("/")[6]).split(" ")[1] + ".xlsx")

                    print("Progress: " + str(counter) + " / " + str(scan_count) + ", " + str(round(((counter / scan_count) * 100), 1)) + " %")

selection = input("Surface(1) selection or Element(2) selection: ")
if selection == "1":
    surface_selection()
elif selection == "2":
    element_selection()                     